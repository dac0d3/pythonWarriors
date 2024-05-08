from tkinter import *
import openpyxl
from openpyxl import Workbook, load_workbook
import customtkinter

from main import Main2

# excel sheet for transactions
book = load_workbook('customerTransactions.xlsx')
sheet = book.active



'''
Programmer's name: Adrian Ramos

Class Description: This class loops through the customer transactions excel sheet and puts all the orders into a list for reference.
The same logic is used to make buttons that associate with a four pizza set. Once clicked a label will be showed with the amount of pizza
for each combination. This class also handles the GUI for the chef window.

Important Functions: 
The getOrders() function is the key to this class. It provides the data that will be used to create the buttons that 
tell the chef what orders to cook. The for loop will place the int values to the CP, PP, HP, and MLP variables which are associated with
a specific type of pizza. 
There is also a for loop in def __init__() that creates button and it's directly associate with the customers order
by matching the index of the orders created in the getOrders() function. The self.order_labels.append(btn) handles this responsibility
The display_row() functions allows the order GUI to disappear and reappear as needed. It also handles the task of displaying the correct order
depending on the button that is choosen. There is an if-else statment that handles this by comparing the text that is associated with the
button. If it matches then the order will disappear on the second click, if it's not, then the new text will display as it is either a 
diifferent order, or identical to the last order meaning no change is needed.

Data Structure(s): Lists

Algorithm: String Matching: The display_row() functon uses this algorithm to detect wether a change needs to be made to the order 
associated with the button or to make the current order disappear if it's no longer needed.

'''



class Chef(Toplevel):

    def __init__(self, parent):
        super().__init__(parent)

        # GUI for chef window
        self.title('Orders')
        self.geometry('1000x500')
        self.config(bg='#d9472a')

        # Canvas to hold the frame and allow scrolling
        self.canvas = Canvas(self, bg='#d9472a')
        self.canvas.pack(side=LEFT, fill=BOTH, expand=True)

        # Canvas to hold the buttons
        self.frame = Frame(self.canvas, bg='#d9472a')
        self.canvas.create_window((0, 0), window=self.frame, anchor='nw')

       # List to hold order buttons and their associated delete buttons
        self.order_buttons = {}

        # list for order data
        self.order_data = []

        # Fetch order data from Excel
        self.getOrders()

        for idx, order_text in enumerate(self.order_data, start=1):
            order_btn = Button(self.frame, text=f"Order {idx}", command=lambda i=idx: self.display_row(i),
                               bg='#d9472a', fg='black', highlightbackground='#d9472a', highlightcolor='#d9472a')
            order_btn.pack(side=TOP, anchor=W, padx=10, pady=5)

            delete_btn = Button(self.frame, text=f"Complete Order {idx}", command=lambda button=order_btn: self.deleteButton(button),
                     bg='#d9472a', fg='black', highlightbackground='#d9472a', highlightcolor='#d9472a')

            delete_btn.pack(side=TOP, anchor=E, padx=10, pady=5)

            # Store order button and delete button in the dictionary
            self.order_buttons[order_btn] = delete_btn
        '''
        # Scrollbar
        scrollbar = Scrollbar(self, orient=VERTICAL, command=self.canvas.yview, bg='blue')
        scrollbar.pack(side=RIGHT, fill=Y)
        self.canvas.config(yscrollcommand=scrollbar.set)

        # Update the scroll function
        self.canvas.update_idletasks()
        self.canvas.config(scrollregion=self.canvas.bbox("all"))
        '''
        # Selected order label
        self.selected_order_label = Label(self, text="", bg='#d9472a')
        self.selected_order_label.pack()

        self.exitButton = customtkinter.CTkButton(self, text='Exit', command=self.exitPage)
        self.exitButton.place(relx=0.8, rely=0.8, anchor=CENTER)

    # Method to fetch orders data from Excel
    def getOrders(self):
        for row in range(2, sheet.max_row + 1):
            CP = sheet['D' + str(row)].value
            PP = sheet['E' + str(row)].value
            HP = sheet['F' + str(row)].value
            MLP = sheet['G' + str(row)].value

            if CP is not None:  # Check if data is available for this row
                self.order_data.append(
                    f"Cheese Pizza(s) {CP}\nPepperoni Pizza(s): {PP}\nHawaiian Pizza(s): {HP}\nMeat Lovers Pizza(s): {MLP}")

    # Method to display selected order details
    def display_row(self, order_num):
        current_text = self.selected_order_label.cget("text")
        new_text = f"ORDER #{order_num}\n{self.order_data[order_num - 1]}" if 1 <= order_num <= len(
            self.order_data) else ""

        if current_text == new_text:
            self.selected_order_label.config(text="", bg='#FFC902', fg='black', font=('Arial', 50))
        else:
            self.selected_order_label.config(text=new_text, bg='#FFC902', fg='black', font=('Arial', 50))

  
    # Method to delete an order button and its associated delete button
    def deleteButton(self, order_btn):
        delete_btn = self.order_buttons.pop(order_btn)
        order_btn.destroy()
        delete_btn.destroy()

    # Check if there are no more order buttons left
        if not self.order_buttons:
            self.selected_order_label.config(text="", bg='#d9472a', fg='black', font=('Arial', 50))


    # Method to exit chef home Page
    def exitPage(self):
        Main2(self)
        self.forget(self)

